import os
import openpyxl
from radiomics import featureextractor

src_path = "..."
# src_path = ".\\data"
dst_path = ".\\fined_data"
MODALITY = ["ADC", "DWI", "T1", "T1C", "T2"]

def extract(scan, label):
    extractor = featureextractor.RadiomicsFeatureExtractor()
    extractor.enableImageTypeByName('Wavelet')
    extractor.enableImageTypeByName('LoG', customArgs={'sigma':[3.0]})
    extractor.enableImageTypeByName('Square')
    extractor.enableImageTypeByName('SquareRoot')
    extractor.enableImageTypeByName('Exponential')
    extractor.enableImageTypeByName('Logarithm')
    extractor.disableAllFeatures()
    extractor.enableFeatureClassByName('firstorder')
    extractor.enableFeatureClassByName('shape')
    extractor.enableFeatureClassByName('glcm')
    extractor.enableFeatureClassByName('glrlm')
    extractor.enableFeatureClassByName('glszm')
    extractor.enableFeatureClassByName('gldm')
    extractor.enableFeatureClassByName('ngtdm')
    feature = extractor.execute(scan, label)
    return feature

def scan_file(directory , prefix=None, postfix=None):
    files_list = []
    for root, sub_dirs, files in os.walk(directory):
        for file in files:
            if postfix:
                if file.endswith(postfix):
                    files_list.append(os.path.join(root, file))
            elif prefix:
                if file.startswith(prefix):
                    files_list.append(os.path.join(root, file))
            else:
                files_list.append(os.path.join(root,file))
    return files_list

def save(data, dst_file, sub_name, modality_name):
    workbook = openpyxl.load_workbook(dst_file)
    worksheet = workbook.active

    write_flag = 0
    index = 1
    current_row = worksheet.max_row+1

    worksheet.cell(row=current_row, column=index).value = sub_name
    for featureName in data.keys():
        write_flag += 1
        if write_flag > 22:
            index += 1
            worksheet.cell(row=current_row, column=index).value = str(data[featureName])
            worksheet.cell(row=1, column=index).value = featureName
    workbook.save(dst_file)

if __name__ == '__main__':
    sub_list = os.listdir(src_path)
    for i, sub in enumerate(sub_list):
        print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++\n")
        print(i+1, "  extracting radiomics features from : ", sub)
        sub_source_path = os.path.join(src_path, sub)
        label_ = scan_file(sub_source_path)[0]
        print(" current label : {}\n".format(label_))

        for j, modality in enumerate(MODALITY):
            print(j+1, "  current modality : ", modality)
            modality_source_path = os.path.join(sub_source_path, modality)
            dst_file = scan_file(dst_path, postfix=modality+".xlsx")[0]
            obj_file = scan_file(modality_source_path)[0]
            featureVector = extract(obj_file, label_)

            save(featureVector, dst_file, sub, modality)
            for featureName in featureVector.keys():
                print("Computed %s : %s" % (featureName, featureVector[featureName]))