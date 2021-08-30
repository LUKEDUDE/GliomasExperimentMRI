import os
import openpyxl

src_dir = ".\\data"
dst_xlsx_path = "..."

if __name__ == '__main__':
    xlsx_list = os.listdir(src_dir)
    for file_name in xlsx_list:
        src_xlsx_path = os.path.join(src_dir, file_name)

        workbook_src = openpyxl.load_workbook(src_xlsx_path)
        worksheet_src = workbook_src.active
        workbook_dst = openpyxl.load_workbook(dst_xlsx_path)
        worksheet_dst = workbook_dst.active

        selected_row = []
        for i in range(2, worksheet_dst.max_row+1):
            for j in range(2, worksheet_src.max_row+1):
                if worksheet_src.cell(row=j, column=1).value == worksheet_dst.cell(row=i, column=1).value:
                    selected_row.append(j)
                    break
        # selected_col = []
        modality_name = worksheet_src.title
        features_ = []
        for j in range(2, worksheet_src.max_column + 1):
            value = worksheet_src.cell(row=1, column=j).value
            str = modality_name + '_' + value
            features_.append(str)

        max_c = worksheet_dst.max_column
        print(max_c)
        for i in range(len(features_)):
            worksheet_dst.cell(row=1, column=max_c+1+i).value = features_[i]
            for j in range(len(selected_row)):
                worksheet_dst.cell(row=j+2, column=max_c+i+1).value = worksheet_src.cell(row=selected_row[j], column=2+i).value
        workbook_dst.save(dst_xlsx_path)