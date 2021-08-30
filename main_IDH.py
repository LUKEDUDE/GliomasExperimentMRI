import openpyxl as opx
import numpy as np
import scipy.stats as stats
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.model_selection import train_test_split, StratifiedKFold
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.feature_selection import SelectFromModel, RFECV
from sklearn.linear_model import ElasticNetCV, SGDClassifier, LassoCV
from sklearn.metrics import  classification_report, confusion_matrix, plot_roc_curve, auc

src_file = ".\\src_data\\subject_list_info_IDH_125.xlsx"
log_DataPrepare_ = ".\\log\\dp_log_.txt"
log_ModelResult_ = ".\\log\\md_log_.txt"
log_result_ = ".\\log\\RandomGroupingExperimentResultLog_.xlsx"
log_pic_ = ".\\log\\myplot_roc.png"

def read_original_data(src):

    workbook = opx.load_workbook(src)
    worksheet = workbook.active

    table = np.zeros((worksheet.max_row-1, worksheet.max_column-1))
    for i in range(2, worksheet.max_row+1):
        for j in range(2, worksheet.max_column+1):
            if worksheet.cell(row=i, column=j).value != None:
                table[i-2, j-2] = worksheet.cell(row=i, column=j).value
    table_X = np.array(table[:, 2:table.shape[1]])
    table_y = np.array(table[:, 0].astype(int))
    table_div = np.array(table[:, 1].astype(int))

    return table_X, table_y

# we need data split and selected features info here
def data_prepare(src):

    #############################################
    #    preparation :
    #############################################

    flog = open(log_DataPrepare_, 'a')
    flog.write("++++++++++++++ data preparation ++++++++++++++++\n")
    workbook = opx.load_workbook(src)
    worksheet = workbook.active

    flog.write("start reading data from source file : {}\n".format(src))
    table_X, table_y = read_original_data(src)
    flog.write("-original- matrix X : {}, label y : {}\n".format(table_X.shape, table_y.shape))

    # default : Grade 0.2
    X_train, X_test, y_train, y_test = train_test_split(
        table_X, table_y, test_size=0.2, shuffle=True
    )
    flog.write("-splited X-  matrix X_train: {}, X_test: {}\n".format(X_train.shape, X_test.shape))
    flog.write("-splited y-  matrix y_train: {}, y_test: {}\n".format(y_train.shape, y_test.shape))

    scaler = StandardScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.transform(X_test)
    encoder = LabelEncoder()
    y_train = encoder.fit_transform(y_train)
    y_test = encoder.transform(y_test)

    ##############################################################
    #       feature selection : Univariate Feature Selection
    ##############################################################
    flog.write("\nRunning Univariate Feature Selection......\n")

    # Step 1 : MannwhitneyU
    selected_byMann = []
    flog.write("\nMannwhitneyU Test: \n\n")
    group_1 = []
    group_2 = []
    for i in range(0, y_train.shape[0]):
        value = y_train[i]
        if value != None:
            if value == 1:
                group_1.append(i+1)
            else:
                group_2.append(i+1)
    flog.write("{} {} : according to the label, cut group into ...\n".format(len(group_1), len(group_2)))

    counter = 0
    for j in range(0, X_train.shape[1]):
        group_1_feature = []
        group_2_feature = []
        for i1 in group_1:
            group_1_feature.append(X_train[i1-1, j])
        for i2 in group_2:
            group_2_feature.append(X_train[i2-1, j])
        try:
            state, p = stats.mannwhitneyu(group_1_feature, group_2_feature, alternative='two-sided')
        except ValueError:
            state, p = (0, 1)
        if p < 0.05:    # default : Grade 0.05
            flog.write(str(worksheet.cell(row=1, column=j+4).value)+"-------------Passed!-------------"+str(p)+"\n")
            selected_byMann.append(1)
            counter+=1
        else:
            flog.write(str(worksheet.cell(row=1, column=j+4).value)+"-------------GivedUp!--------------"+str(p)+"\n")
            selected_byMann.append(0)

    # Step 2 : select from model ( lasso or elastic net )
    flog.write("\nselected from model: \n\n")
    table_temp = np.zeros((y_train.shape[0], counter))
    for i in range(0, y_train.shape[0]):
        c = 0
        for j in range(0, X_train.shape[1]):
            temp_flag = True if selected_byMann[j] else False
            if temp_flag:
                table_temp[i, c] = X_train[i, j]
                c += 1
    flog.write(
        "build the matrix X after MannwhitneyU test, prepare for the SFM.\n{}\n{}\n".format(str(table_temp.shape),
                                                                                            str(y_train.shape))
    )

    sfm_model = ElasticNetCV().fit(table_temp, y_train)     # default : Grade elasticnet
    model = SelectFromModel(sfm_model, prefit=True)
    X_train_US = model.transform(table_temp)
    selected_bySFM = model.get_support()
    flog.write(
        "Finally, there's {} features selected in {} after SFM.\n".format(str(X_train_US.shape[1]),
                                                                          str(table_temp.shape[1]))
    )

    # Step 3 : ...
    selected_ = []
    temp = 0
    for i in range(len(selected_byMann)):
        if selected_byMann[i]:
            if selected_bySFM[temp]:
                selected_.append(1)
            else:
                selected_.append(0)
            temp += 1
        else:
            selected_.append(0)

    fn_ = []
    flog.write("\nGet the features' name after the US.\n")
    for j in range(4, worksheet.max_column+1):
        if selected_[j-4]:
            feature_name = worksheet.cell(row=1, column=j).value
            fn_.append(feature_name)
            flog.write(str(feature_name)+"\n")

    # Step 4 : build train & test matrix
    y_test_US = y_test
    y_train_US = y_train
    X_test_US = np.zeros((y_test_US.shape[0], X_train_US.shape[1]))
    c = 0
    for j in range(0, X_test.shape[1]):
        if selected_[j]:
            for i in range(0, y_test_US.shape[0]):
                X_test_US[i, c] = X_test[i, j]
            c += 1
        else:
             continue

    # current result : X_train_US, X_test_US, y_train_US, y_test_US, fn_

    ##############################################################
    #       cut off redundant features : collinearity test & rfecv
    ##############################################################
    flog.write("\nRunning redundant features detection......\n")

    # Step 1 : collinearity test
    flog.write("colinearity test: \n\n")
    X_train_temp = pd.DataFrame(X_train_US, columns=fn_)
    corrmat = X_train_temp.corr()

    threshold = 1.0     # default : Grade 1.0
    deleted_ = np.zeros((X_train_US.shape[1]))
    counter_del_ = 0
    for j in range(0, corrmat.shape[1]):
        temp_flag = 0
        for i in range(0, corrmat.shape[0]):
            if abs(corrmat.iat[i, j]) >= threshold :
                if deleted_[i] != 0:
                    continue
                elif deleted_[i] == 0 and j!=i:
                    temp_flag = 1
                    break
        if temp_flag == 1:
            deleted_[j] = 1
            counter_del_ += 1

    X_train_ct = np.zeros((X_train_US.shape[0], X_train_US.shape[1]-counter_del_))
    X_test_ct = np.zeros((X_test_US.shape[0], X_test_US.shape[1]-counter_del_))
    fn_ct_ = []
    temp = 0
    for j in range(0, X_train_US.shape[1]):
        if deleted_[j] != 1:
            for i in range(0, X_train_US.shape[0]):
                X_train_ct[i, temp] = X_train_US[i, j]
            for i in range(0, X_test_US.shape[0]):
                X_test_ct[i, temp] = X_test_US[i, j]
            fn_ct_.append(fn_[j])
            temp += 1

    # X_train_temp = pd.DataFrame(X_train_ct, columns=fn_ct_)
    # corrmat = X_train_temp.corr()
    # f, ax = plt.subplots(figsize=(12, 9))
    # sns.heatmap(corrmat, square=True, cmap='RdBu_r', vmax=1.0, vmin=-1.0)
    # plt.show()

    flog.write("\nthe feature num after ct : {}\n\n".format(X_train_ct.shape[1]))
    for i in range(len(fn_ct_)):
        flog.write(str(fn_ct_[i])+'\n')

    # ct result : X_train_ct, X_test_ct, y_train_US, y_test_US, fn_ct_

    # Step 2 : rfecv
    flog.write("\nrfecv : \n")

    model = SGDClassifier(loss='hinge')
    min_features_to_select = 1
    rfecv = RFECV(estimator=model, step=1, cv=StratifiedKFold(2),
                  scoring='accuracy', min_features_to_select=min_features_to_select)
    rfecv.fit(X_train_ct, y_train_US)
    flog.write("\nOptimal number of features : {}\n".format(rfecv.n_features_))

    X_train_rfe = np.zeros((X_train_ct.shape[0], rfecv.n_features_))
    X_test_rfe = np.zeros((X_test_ct.shape[0], rfecv.n_features_))
    temp = 0
    fn_rfe_ = []
    for j in range(0, X_train_ct.shape[1]):
        if rfecv.support_[j]:
            for i in range(0, X_train_ct.shape[0]):
                X_train_rfe[i, temp] = X_train_ct[i, j]
            for i in range(0, X_test_ct.shape[0]):
                X_test_rfe[i, temp] = X_test_ct[i, j]
            fn_rfe_.append(fn_ct_[j])
            temp += 1

    # plt.figure()
    # plt.xlabel("Number of features selected")
    # plt.ylabel("Cross Validation Score (nb of correct classification)")
    # plt.plot(range(min_features_to_select,
    #                len(rfecv.grid_scores_) + min_features_to_select),
    #          rfecv.grid_scores_)
    # plt.show()

    flog.write("\n the feature num after rfecv : {}\n\n".format(rfecv.n_features_))
    for i in range(len(fn_rfe_)):
        flog.write(str(fn_rfe_[i]+'\n'))
    flog.close()

    # rfecv result : X_train_rfe, X_test_rfe, y_train_US, y_test_US, fn_rfe_

    return fn_rfe_, X_train_rfe, X_test_rfe, y_train_US, y_test_US

# perform permutation test to valid classification
def permutation_test(test_acc, X_train, X_test, y_train, y_test):

    acc_list = []
    for f in range(0, 1000):
        print("\n {} +++++++++++++++++++++++++++++++++++++++++++\n".format(f))
        np.random.shuffle(y_train)
        np.random.shuffle(y_test)

        model = SGDClassifier(loss='hinge')
        model.fit(X_train, y_train)

        acc = model.score(X_test, y_test)
        acc_list.append(acc)

    acc_list.sort(reverse=True)
    counter = 0
    for i in range(0, len(acc_list)):
        counter += 1
        if acc_list[i] <= test_acc:
            break

    return counter/1000

def act_ClassifyTest(src):

    fn_, X_train, X_test, y_train, y_test = data_prepare(src)
    flog = open(log_ModelResult_, 'a')

    # act SGD Classify process next
    model = SGDClassifier(loss='hinge')
    model.fit(X_train, y_train)

    # evaluation
    flog.write("++++++++++++++++++++++++++++++++++++++++++++")

    # classification report :
    flog.write("\n\nDetailed classification report : \n")
    y_true, y_pred = y_test, model.predict(X_test)
    flog.write(classification_report(y_true, y_pred))

    # sensitivity & specificity & precision & f1-score
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred).ravel()
    sensitivity = tp/(tp+fn)
    specificity = tn/(fp+tn)
    precision = tp/(tp+fp)
    f1score = 2*((precision*sensitivity)/(precision+sensitivity))
    flog.write("\nsensitivity & specificity : {}  {}\n".format(sensitivity, specificity))
    flog.write("precision & f1score : {}  {}\n".format(precision, f1score))

    # training & testing accuracy
    train_acc = model.score(X_train, y_train)
    test_acc = model.score(X_test, y_test)
    flog.write("\nfinal result : \n")
    flog.write("score on train set : {}\n".format(train_acc))
    flog.write("score on test set : {}\n".format(test_acc))

    # ROC & AUC
    plt.figure()
    roc = plot_roc_curve(model, X_test, y_test)
    plt.plot([0, 1], [0, 1], color='navy', linestyle='--')
    plt.show()
    auc = roc.roc_auc
    flog.write("\nauc on test set : {}\n".format(auc))

    # feature_num
    fea_num = len(fn_)
    flog.write("\nfinal feature number : {}\n".format(fea_num))
    #
    for i in range(fea_num):
        flog.write(fn_[i]+"\n")

    # permuation test : p_value
    p_value = permutation_test(test_acc, X_train, X_test, y_train, y_test)
    flog.write("\n permutation test p_value : {}\n".format(p_value))

    flog.close()

    return roc.fpr, roc.tpr, auc, \
           [auc, train_acc, test_acc, sensitivity, specificity, precision, f1score, fea_num, p_value]

def grouping_roc(iter, fprs_, tprs_, aucs_):

    tprs = []
    aucs = []
    mean_fpr = np.linspace(0, 1, 100)
    fig, ax = plt.subplots()
    print(fprs_, tprs_, aucs_)

    for index in range(iter):
        fpr = fprs_[index]
        tpr = tprs_[index]
        roc_auc = aucs_[index]
        print(fpr, tpr, roc_auc)

        ax.plot(fpr, tpr, label='ROC group %d (AUC = %0.2f)'%(index+1, roc_auc), alpha=0.3, lw=1)

        interp_tpr = np.interp(mean_fpr, fpr, tpr)
        interp_tpr[0] = 0.0
        tprs.append(interp_tpr)
        aucs.append(roc_auc)

    ax.plot([0, 1], [0, 1], linestyle='--', lw=2, color='r', label='Chance', alpha=.8)

    mean_tpr = np.mean(tprs, axis=0)
    mean_tpr[-1] = 1.0
    mean_auc = auc(mean_fpr, mean_tpr)
    std_auc = np.std(aucs)
    ax.plot(mean_fpr, mean_tpr, color='b',
            label=r'Mean ROC (AUC = %0.2f $\pm$ %0.2f)'%(mean_auc, std_auc),
            lw=2, alpha=.8)

    std_tpr = np.std(tprs, axis=0)
    tprs_upper = np.minimum(mean_tpr + std_tpr, 1)
    tprs_lower = np.maximum(mean_tpr - std_tpr, 0)
    ax.fill_between(mean_fpr, tprs_lower, tprs_upper, color='grey', alpha=.2,
                    label=r'$\pm$ 1 std. dev.')

    ax.set(xlim=[-0.05, 1.05], ylim=[-0.05, 1.05], title="Receiver operating characteristic example")
    ax.legend(loc="lower right")
    plt.savefig(log_pic_)
    plt.show()

    return

if __name__ == '__main__':

    maximum_num_of_exp = 30     # random grouping experiment samples
    index = 0                   # current index
    experiment_result = np.zeros((maximum_num_of_exp, 9))

    fpr_list = []
    tpr_list = []
    auc_list = []

    for index in range(maximum_num_of_exp):

        print(" index: {} \n".format(index+1))
        fpr, tpr, roc_auc, result_ = act_ClassifyTest(src_file)        # return with a result group of current random grouping test
        experiment_result[index, :] = result_

        fpr_list.append(fpr)
        tpr_list.append(tpr)
        auc_list.append(roc_auc)

    grouping_roc(maximum_num_of_exp, fpr_list, tpr_list, auc_list)

    #
    # output the random grouping experiment results
    # and calculate the mean + std value / 95%CI for : AUC ACC(train) ACC(test) Sensitivity Specificity
    #                                               Precision F1Score feature_num p_value
    #

    workbook = opx.load_workbook(log_result_)
    worksheet = workbook.active

    for j in range(0, experiment_result.shape[1]):

        temp = np.array(experiment_result[:, j])
        mean = temp.mean()
        std = temp.std()
        a = mean - 1.96*std/np.sqrt(maximum_num_of_exp)
        b = mean + 1.96*std/np.sqrt(maximum_num_of_exp)
        # interval = stats.t.interval(0.95, len(temp)-1, mean, std)       # 95%CI

        for i in range(0, maximum_num_of_exp):
            worksheet.cell(row=i+2, column=j+2).value = temp[i]

        worksheet.cell(row=maximum_num_of_exp+2, column=j+2).value = mean
        worksheet.cell(row=maximum_num_of_exp+3, column=j+2).value = std
        worksheet.cell(row=maximum_num_of_exp+4, column=j+2).value = a
        worksheet.cell(row=maximum_num_of_exp+5, column=j+2).value = b

    workbook.save(log_result_)